import tkinter as tk
from tkinter import filedialog, messagebox, IntVar, Checkbutton
from tkinter.ttk import Progressbar
import pandas as pd
from openpyxl import load_workbook
from datetime import datetime, timedelta

# 设置试用期结束日期（精确到时分秒）
# trial_end_datetime = datetime(2025, 3, 31, 20, 00, 00)  # 试用期结束时间

# 计算剩余试用时间（精确到时分秒）
# def get_remaining_trial_time():
#     now = datetime.now()
#     if now > trial_end_datetime:
#         return "试用期已过"  # 如果试用期已过
#     else:
#         remaining_time = trial_end_datetime - now
#         hours, remainder = divmod(remaining_time.seconds, 3600)
#         minutes, seconds = divmod(remainder, 60)
#         return f"{remaining_time.days}天 {hours}小时 {minutes}分钟 {seconds}秒"

# 处理Excel数据的逻辑
def process_excel(source_file, target_file, months, progress_var, progress_label):
    source_data = pd.read_excel(source_file)
    target_data = pd.read_excel(target_file, header=[0, 1, 2])  # 读取前三级作为多级表头

    # 定义表头和月份
    sub_headers = ['时间', '正向有功', '正向无功']
    sub_sub_headers = ['总（kWh）', '尖（kWh）', '峰（kWh）', '平（kWh）', '谷（kWh）', '总（kVarh）']

    # 动态生成 columns_mapping
    columns_mapping = {}
    for month in months:
        columns_mapping[(month, '时间')] = ((int(month.replace('月', '')) - 1) * 7 + 11)

        for sub_header in sub_headers:
            if sub_header == '时间':
                continue

            for sub_sub_header in sub_sub_headers:
                if '正向有功' in sub_header and sub_sub_header in ['总（kWh）', '尖（kWh）', '峰（kWh）', '平（kWh）',
                                                                   '谷（kWh）']:
                    col_idx = ((int(month.replace('月', '')) - 1) * 7 + 12) + sub_sub_headers.index(sub_sub_header)
                    columns_mapping[(month, sub_header, sub_sub_header)] = col_idx

                elif '正向无功' in sub_header and sub_sub_header == '总（kVarh）':
                    col_idx = ((int(month.replace('月', '')) - 1) * 7 + 12) + len(sub_sub_headers) - 1
                    columns_mapping[(month, sub_header, sub_sub_header)] = col_idx

    # 加载目标文件，准备更新数据
    workbook = load_workbook(target_file)
    sheet = workbook.active

    # 遍历源数据，填充到目标表格
    total_rows = len(source_data)
    for index, row in source_data.iterrows():
        progress_var.set(int((index / total_rows) * 100))
        progress_label.config(text=f"正在处理: {index + 1}/{total_rows}")
        root.update_idletasks()

        # match_condition = (
        #         (target_data.iloc[:, 9] == row['终端资产']) |  # 匹配终端编号
        #         (target_data.iloc[:, 10] == row['表资产号'])  # 匹配电能表编号
        # )
        match_condition = (
                (target_data.iloc[:, 10] == row['表资产号'])  # 匹配电能表编号
        )
        target_row_index = target_data.loc[match_condition].index

        if not target_row_index.empty:
            target_row_index = target_row_index + 4
            for col, col_idx in columns_mapping.items():
                header = col[0]
                sub_header = col[1]
                sub_sub_header = col[2] if len(col) > 2 else None

                # 填充数据
                if sub_header == '时间' and sub_sub_header is None:
                    target_cell = sheet.cell(row=target_row_index[0], column=col_idx)
                    target_cell.value = row['数据时间']
                elif sub_header == '正向有功' and sub_sub_header == '总（kWh）':
                    target_cell = sheet.cell(row=target_row_index[0], column=col_idx)
                    target_cell.value = row['正向有功总']
                elif sub_header == '正向有功' and sub_sub_header == '尖（kWh）':
                    target_cell = sheet.cell(row=target_row_index[0], column=col_idx)
                    target_cell.value = row['正向有功尖']
                elif sub_header == '正向有功' and sub_sub_header == '峰（kWh）':
                    target_cell = sheet.cell(row=target_row_index[0], column=col_idx)
                    target_cell.value = row['正向有功峰']
                elif sub_header == '正向有功' and sub_sub_header == '平（kWh）':
                    target_cell = sheet.cell(row=target_row_index[0], column=col_idx)
                    target_cell.value = row['正向有功平']
                elif sub_header == '正向有功' and sub_sub_header == '谷（kWh）':
                    target_cell = sheet.cell(row=target_row_index[0], column=col_idx)
                    target_cell.value = row['正向有功谷']
                elif sub_header == '正向无功' and sub_sub_header == '总（kVarh）':
                    target_cell = sheet.cell(row=target_row_index[0], column=col_idx)
                    target_cell.value = row['正向无功总']

    workbook.save(target_file)
    print(f"数据已成功填充并保存到 {target_file}")
    messagebox.showinfo("完成", "数据处理完成并已保存到目标文件！")


# 修改选择文件的功能，自动选择文件
def select_file(file_label):
    file_path = filedialog.askopenfilename(filetypes=[("Excel Files", "*.xls *.xlsx")])
    if file_path:
        file_label.config(text=f"已选择文件: {file_path}")
        return file_path
    else:
        file_label.config(text="未选择文件")
        return None


def start_processing():
    # 获取选择的文件
    source_file = file_label.cget("text").replace("已选择文件: ", "")
    target_file = target_file_label.cget("text").replace("已选择文件: ", "")

    # 获取选择的月份
    months = [month for month, var in month_vars.items() if var.get() == 1]

    if not months:
        messagebox.showwarning("警告", "请选择至少一个月份！")
        return

    if not source_file or not target_file:
        messagebox.showwarning("警告", "请选择源文件和目标文件！")
        return

    try:
        process_excel(source_file, target_file, months, progress_var, progress_label)
    except Exception as e:
        messagebox.showerror("错误", f"处理文件时出错: {e}")


# 创建 GUI 界面
root = tk.Tk()
root.title("油田电子表数据-Excel文件处理工具")
root.geometry("600x450")

# 文件选择部分
file_label = tk.Label(root, text="点击下方按钮选择 Excel 文件", font=("Arial", 12))
file_label.pack(pady=10)

select_button = tk.Button(root, text="选择源文件", command=lambda: select_file(file_label))
select_button.pack(pady=10)

target_file_label = tk.Label(root, text="点击下方按钮选择目标 Excel 文件", font=("Arial", 12))
target_file_label.pack(pady=10)

select_target_button = tk.Button(root, text="选择目标文件", command=lambda: select_file(target_file_label))
select_target_button.pack(pady=10)

# 月份选择复选框
month_labels = ['1月', '2月', '3月', '4月', '5月', '6月', '7月', '8月', '9月', '10月', '11月', '12月']
month_vars = {}

# 创建一个 Frame 用于包含月份复选框，使用 grid 布局
month_frame = tk.Frame(root)
month_frame.pack(pady=10)

# 使用 grid 布局将月份分为两行显示
for i, month in enumerate(month_labels):
    month_vars[month] = IntVar()
    row = i // 6  # 计算行数，将月份分为两行
    col = i % 6  # 计算列数
    Checkbutton(month_frame, text=month, variable=month_vars[month]).grid(row=row, column=col, pady=2)

# 进度条
progress_var = tk.IntVar()
progress_bar = Progressbar(root, orient="horizontal", length=400, mode="determinate", variable=progress_var)
progress_bar.pack(pady=10)

progress_label = tk.Label(root, text="", font=("Arial", 10), fg="green")
progress_label.pack(pady=5)

# 显示剩余试用时间
# remaining_trial_time = get_remaining_trial_time()
# trial_label = tk.Label(root, text=f"剩余试用时间: {remaining_trial_time}", font=("Arial", 12), fg="red")
# trial_label.pack(pady=10)

# 更新剩余时间，每秒刷新
# def update_remaining_time():
#     remaining_time = get_remaining_trial_time()
#     trial_label.config(text=f"剩余试用时间: {remaining_time}")
#     root.after(1000, update_remaining_time)  # 每秒更新一次

# update_remaining_time()

# 处理按钮
process_button = tk.Button(root, text="开始处理", command=start_processing)
process_button.pack(pady=10)

root.mainloop()
