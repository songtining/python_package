import os
import tkinter as tk
from tkinter import filedialog, messagebox, StringVar, OptionMenu
import pandas as pd
from openpyxl import load_workbook
from openpyxl.utils import get_column_letter
from tkinter.ttk import Progressbar, Style

# 去重处理（按 LAC 和 CI 去重）
def deduplicate(df):
    """按 LAC 和 CI 去重，优先保留经纬度不为 0 的数据"""
    deduplicated = []
    grouped = df.groupby(["CGI（必填，CGI序列或运营商名称）", "LAC（必填）", "CI（必填）"])

    for _, group in grouped:
        # 检查是否存在经纬度不为 0 的行
        valid_rows = group[(group["基站经度"] != 0) & (group["基站纬度"] != 0)]
        if not valid_rows.empty:
            # 如果存在经纬度不为 0 的数据，优先保留第一条
            deduplicated.append(valid_rows.iloc[0])
        else:
            # 如果所有行的经纬度均为 0，随机保留一条
            deduplicated.append(group.iloc[0])

    return pd.DataFrame(deduplicated)


# CGI 对应关系字典
CGI_MAPPING = {
    "ChinaMobileGsm": "移动2G",
    "ChinaMobileTdscdma": "移动3G",
    "ChinaMobileLte": "移动4G",
    "ChinaMobileNR": "移动5G",
    "ChinaUnionGsm": "联通2G",
    "ChinaUnionWcdma": "联通3G",
    "ChinaUnionLte": "联通4G",
    "ChinaUnionNR": "联通5G",
    "ChinaTelecomCdma": "电信2G3G",
    "ChinaTelecomLte": "电信4G",
    "ChinaTelecomNR": "电信5G"
}


def process_excel(input_file, output_file, progress_var, progress_label, output_label):
    # 初始化一个空的 DataFrame 来存储表格2数据
    columns_table2 = ["CGI（必填，CGI序列或运营商名称）", "LAC（必填）", "CI（必填）",
                      "基站类型（必填）", "基站经度", "基站纬度", "基站名称", "基站地址"]
    result_df = pd.DataFrame(columns=columns_table2)

    # 判断文件扩展名以选择合适的 engine
    file_extension = os.path.splitext(input_file)[1].lower()
    engine = "openpyxl" if file_extension == ".xlsx" else "xlrd"

    # 读取表格1中的所有 sheet 页
    sheets = pd.ExcelFile(input_file, engine=engine)
    total_sheets = len(sheets.sheet_names)
    for idx, sheet_name in enumerate(sheets.sheet_names, start=1):
        if sheet_name == 'WIFI':
            continue
        # 更新进度条
        progress_var.set(int((idx / total_sheets) * 100))
        progress_label.config(text=f"正在处理: {sheet_name} ({idx}/{total_sheets})")
        root.update_idletasks()

        # 读取每个 sheet 页
        sheet_df = pd.read_excel(input_file, sheet_name=sheet_name, engine=engine)

        # CGI 值来源于映射关系
        cgi_value = CGI_MAPPING.get(sheet_name, '')
        if cgi_value != '':
            cgi_value = cgi_value[:2]  # 截取映射值前两个字符
        else:
            # 如果映射表中不存在
            if any(keyword in sheet_name for keyword in ["移动", "电信", "联通"]):
                # 保留包含 "移动"、"电信"、"联通" 的部分
                cgi_value = next((keyword for keyword in ["移动", "电信", "联通"] if keyword in sheet_name), sheet_name)
            else:
                cgi_value = sheet_name

        # 检查列的数量是否足够
        if sheet_df.shape[1] < 3:
            print(f"Sheet 页 {sheet_name} 列数量不足，跳过处理。")
            continue

        # 提取 LAC（第二列）和 CI（第三列）
        lac_col = sheet_df.iloc[:, 1]  # 第二列
        ci_col = sheet_df.iloc[:, 2]  # 第三列

        # 提取经纬度列（列名固定为 lng 和 lat）
        if "lng" not in sheet_df.columns or "lat" not in sheet_df.columns:
            print(f"Sheet 页 {sheet_name} 缺少经纬度列，跳过处理。")
            continue

        # 提取经纬度列
        lng_col = sheet_df["lng"]
        lat_col = sheet_df["lat"]

        # 将数据按照规则转换为表格2格式
        transformed_df = pd.DataFrame({
            "CGI（必填，CGI序列或运营商名称）": cgi_value,
            "LAC（必填）": lac_col,
            "CI（必填）": ci_col,
            "基站类型（必填）": "运营商基站",  # 默认填写（可根据需要修改）
            "基站经度": lng_col,
            "基站纬度": lat_col,
            "基站名称": None,  # 没有对应列，填充为空值
            "基站地址": None  # 没有对应列，填充为空值
        })

        # 汇总到结果 DataFrame
        result_df = pd.concat([result_df, transformed_df], ignore_index=True)

    # 去重处理（按 LAC 和 CI 去重）
    result_df = deduplicate(result_df)

    # 移除 LAC 和 CI 都为 0 的无效数据
    result_df = result_df[(result_df["LAC（必填）"] != 0) | (result_df["CI（必填）"] != 0)]

    # 对最终的数据按 LAC（必填） 排序
    result_df = result_df.sort_values(by=["CGI（必填，CGI序列或运营商名称）"], ascending=True)

    # 将经纬度列转为字符串格式，保留原始小数位数
    result_df["基站经度"] = result_df["基站经度"].apply(lambda x: str(x) if pd.notnull(x) else "")
    result_df["基站纬度"] = result_df["基站纬度"].apply(lambda x: str(x) if pd.notnull(x) else "")

    # 写入到本地文件
    result_df.to_excel(output_file, index=False, engine="openpyxl")

    # 设置输出的列宽
    workbook = load_workbook(output_file)
    worksheet = workbook.active

    # 设置每列的宽度为 35
    for i, column in enumerate(columns_table2, start=1):
        col_letter = get_column_letter(i)
        worksheet.column_dimensions[col_letter].width = 35  # 宽度设置为 35

    # 保存文件
    workbook.save(output_file)
    print(f"数据已成功处理并保存到 {output_file}，并设置了列宽。")
    progress_var.set(100)
    progress_label.config(text="处理完成！")
    output_label.config(text=f"文件已保存到: {output_file}")
    messagebox.showinfo("完成", f"处理后的文件已成功处理并保存到 {output_file}")

def process_excel_format2(input_file, output_file, progress_var, progress_label, output_label):
    # 初始化一个空的 DataFrame 来存储表格2数据
    columns_table2 = ["CGI（必填，CGI序列或运营商名称）", "LAC（必填）", "CI（必填）",
                      "基站类型（必填）", "基站经度", "基站纬度", "基站名称", "基站地址"]
    result_df = pd.DataFrame(columns=columns_table2)

    # 判断文件扩展名以选择合适的 engine
    file_extension = os.path.splitext(input_file)[1].lower()
    engine = "openpyxl" if file_extension == ".xlsx" else "xlrd"

    # 使用 pandas 读取 Excel 文件，指定前两行作为表头
    df = pd.read_excel(input_file, header=[0, 1], engine=engine)  # 如果是 .xlsx 文件
    # 扁平化多级表头为单级表头，组合第一行和第二行
    df.columns = ['_'.join(col).strip() for col in df.columns.values]

    # 获取所有一级表头
    first_level_headers = set([col.split('_')[0] for col in df.columns])

    # 筛选出包含 "电信"、"移动"、"联通" 和 "坐标" 的一级表头
    valid_headers = {header for header in first_level_headers if
                     any(keyword in header for keyword in ["电信", "移动", "联通", "坐标"])}

    # 获取坐标表头相关的数据
    coordinates_header = [header for header in valid_headers if "坐标" in header]
    if coordinates_header:
        coordinates_header = coordinates_header[0]
        # 获取坐标相关的列（包含经度和纬度的列）
        longitude_column = [col for col in df.columns if coordinates_header in col and "经度" in col]
        latitude_column = [col for col in df.columns if coordinates_header in col and "纬度" in col]

    # 用于存储结果的列表
    merged_data = []
    # 打印每个一级表头对应的数据
    for header in valid_headers:
        # 如果表头是坐标，则跳过
        if "坐标" in header:
            continue

        print(f"\n数据对应表头: {header}")
        # 筛选该一级表头对应的所有列
        selected_columns = [col for col in df.columns if col.startswith(header)]

        # 筛选含 "AC" 或 "CI" 的列
        lac_column = [col for col in selected_columns if 'AC' in col or 'CI' in col]
        ci_column = [col for col in selected_columns if 'CI' in col]

        # 确保每个一级表头包含 LAC 和 CI 的列
        if lac_column and ci_column:
            # 提取LAC和CI的列数据
            lac_data = df[lac_column[0]]
            ci_data = df[ci_column[0]]

            # 获取坐标表头的经度和纬度列数据
            if longitude_column and latitude_column:
                longitude_data = df[longitude_column[0]]
                latitude_data = df[latitude_column[0]]

            # 创建新的行数据
            for i in range(len(lac_data)):
                merged_row = {
                    'CGI（必填，CGI序列或运营商名称）': header[:2],
                    'LAC（必填）': lac_data.iloc[i],
                    'CI（必填）': ci_data.iloc[i],
                    '基站类型': '运营商基站',  # 固定值
                    '基站经度': longitude_data.iloc[i] if longitude_column else None,
                    '基站纬度': latitude_data.iloc[i] if latitude_column else None
                }
                merged_data.append(merged_row)

    # 转换成DataFrame
    merged_df = pd.DataFrame(merged_data)

    # 去重处理（按 LAC 和 CI 去重）
    def deduplicate(df):
        """按 LAC 和 CI 去重，优先保留经纬度不为 0 的数据"""
        deduplicated = []
        grouped = df.groupby(["CGI（必填，CGI序列或运营商名称）", "LAC（必填）", "CI（必填）"])

        for _, group in grouped:
            # 检查是否存在经纬度不为 0 的行
            valid_rows = group[(group["基站经度"] != 0) & (group["基站纬度"] != 0)]
            if not valid_rows.empty:
                # 如果存在经纬度不为 0 的数据，优先保留第一条
                deduplicated.append(valid_rows.iloc[0])
            else:
                # 如果所有行的经纬度均为 0，随机保留一条
                deduplicated.append(group.iloc[0])

        return pd.DataFrame(deduplicated)

    # 去重处理（按 LAC 和 CI 去重）
    merged_df = deduplicate(merged_df)

    # 移除 LAC 和 CI 都为 0 的无效数据
    merged_df = merged_df[(merged_df["LAC（必填）"] != 0) | (merged_df["CI（必填）"] != 0)]

    # 对最终的数据按 LAC（必填） 排序
    merged_df = merged_df.sort_values(by=["CGI（必填，CGI序列或运营商名称）"], ascending=True)

    # 将经纬度列转为字符串格式，保留原始小数位数
    merged_df["基站经度"] = merged_df["基站经度"].apply(lambda x: str(x) if pd.notnull(x) else "")
    merged_df["基站纬度"] = merged_df["基站纬度"].apply(lambda x: str(x) if pd.notnull(x) else "")

    result_df = merged_df.copy()
    # 写入到本地文件
    result_df.to_excel(output_file, index=False, engine="openpyxl")

    # 设置输出的列宽
    workbook = load_workbook(output_file)
    worksheet = workbook.active

    # 设置每列的宽度为 35
    for i, column in enumerate(columns_table2, start=1):
        col_letter = get_column_letter(i)
        worksheet.column_dimensions[col_letter].width = 35  # 宽度设置为 35

    # 保存文件
    workbook.save(output_file)
    print(f"数据已成功处理并保存到 {output_file}，并设置了列宽。")
    progress_var.set(100)
    progress_label.config(text="处理完成！")
    output_label.config(text=f"文件已保存到: {output_file}")
    messagebox.showinfo("完成", f"处理后的文件已成功处理并保存到 {output_file}")

def select_file():
    # 修改文件选择对话框，支持 .xls 和 .xlsx 格式
    file_path = filedialog.askopenfilename(filetypes=[("Excel Files", "*.xls *.xlsx")])
    if file_path:
        file_label.config(text=f"已选择文件: {file_path}")
        process_button.config(state=tk.NORMAL)
        return file_path
    else:
        file_label.config(text="未选择文件")
        process_button.config(state=tk.DISABLED)


def start_processing():
    input_file = file_label.cget("text").replace("已选择文件: ", "")
    if not input_file or not os.path.isfile(input_file):
        messagebox.showwarning("警告", "未选择有效的输入文件！")
        return

    if file_type_var == "格式1(多页)":
        # 输出文件路径：与输入文件同级目录，文件名格式为“输入文件名_基站信息导入.xlsx”
        output_file = os.path.join(
            os.path.dirname(input_file),
            f"{os.path.splitext(os.path.basename(input_file))[0]}_基站信息导入_格式1.xlsx"
        )

        try:
            process_excel(input_file, output_file, progress_var, progress_label, output_label)
        except Exception as e:
            messagebox.showerror("错误", f"处理文件时出错: {e}")
    elif file_type_var == "格式2(多级表头)":
        # 输出文件路径：与输入文件同级目录，文件名格式为“输入文件名_基站信息导入.xlsx”
        output_file = os.path.join(
            os.path.dirname(input_file),
            f"{os.path.splitext(os.path.basename(input_file))[0]}_基站信息导入_格式2.xlsx"
        )
        try:
            process_excel_format2(input_file, output_file, progress_var, progress_label, output_label)
        except Exception as e:
            messagebox.showerror("错误", f"处理文件时出错: {e}")


# 创建 GUI 界面
root = tk.Tk()
root.title("Excel文件处理工具")
root.geometry("600x400")
root.resizable(False, False)

# 样式设置
style = Style()
style.configure("TProgressbar", thickness=15)

# 文件选择部分
file_label = tk.Label(root, text="点击下方按钮选择Excel文件", font=("Arial", 12))
file_label.pack(pady=10)

select_button = tk.Button(root, text="选择文件", command=select_file)
select_button.pack(pady=10)

# 文件选择部分
file_label = tk.Label(root, text="请选择输入的Excel文件数据格式", font=("Arial", 12))
file_label.pack(pady=10)
# 文件类型选择
file_type_var = StringVar(root)
file_type_var.set("格式1(多页)")  # 默认选择“格式1”
file_type_menu = OptionMenu(root, file_type_var, "格式1(多页)", "格式2(多级表头)")
file_type_menu.pack(pady=20)

# 进度条
progress_var = tk.IntVar()
progress_bar = Progressbar(root, orient="horizontal", length=400, mode="determinate", variable=progress_var)
progress_bar.pack(pady=10)

progress_label = tk.Label(root, text="", font=("Arial", 10), fg="green")
progress_label.pack(pady=5)

# 输出文件路径显示
output_label = tk.Label(root, text="", font=("Arial", 10), fg="blue", wraplength=500, justify="center")
output_label.pack(pady=10)

# 处理按钮
process_button = tk.Button(root, text="开始处理", state=tk.DISABLED, command=start_processing)
process_button.pack(pady=10)

# 主循环
root.mainloop()
